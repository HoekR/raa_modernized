"""Tests for editorial Excel/CSV import."""

from __future__ import annotations

import csv
import io

from openpyxl import Workbook

from raa_api.editorial_import import (
    IMPORT_COLUMN_KEYS,
    WORKSHEET_NAME,
    build_persoon_template_xlsx,
    parse_persoon_import_file,
)


def _csv_bytes(rows: list[dict[str, str]]) -> bytes:
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=list(IMPORT_COLUMN_KEYS))
    writer.writeheader()
    writer.writerows(rows)
    return buf.getvalue().encode("utf-8-sig")


def test_template_has_fixed_headers():
    data = build_persoon_template_xlsx()
    assert len(data) > 1000
    parsed = parse_persoon_import_file("test.xlsx", data)
    assert parsed["rows_parsed"] == 0


def test_parse_csv_row():
    payload = _csv_bytes(
        [
            {
                "persoon_id": "42",
                "geslachtsnaam": "Janssen",
                "voornaam": "",
                "tussenvoegsel": "",
                "geboorte_j": "1701",
                "geboorte_m": "",
                "geboorte_d": "",
                "overlijden_j": "",
                "overlijden_m": "",
                "overlijden_d": "",
                "opmerkingen": "",
            }
        ]
    )
    parsed = parse_persoon_import_file("werklijst.csv", payload)
    assert not parsed["parse_errors"]
    assert parsed["change_count"] == 2
    assert parsed["changes"][0]["field"] == "geslachtsnaam"


def test_reject_bad_headers():
    payload = b"id,naam\n1,x\n"
    parsed = parse_persoon_import_file("bad.csv", payload)
    assert parsed["parse_errors"]
    assert "Kolomkoppen" in parsed["parse_errors"][0]["error"]


def test_clear_marker():
    payload = _csv_bytes(
        [
            {
                "persoon_id": "7",
                "geslachtsnaam": "-",
                "voornaam": "",
                "tussenvoegsel": "",
                "geboorte_j": "",
                "geboorte_m": "",
                "geboorte_d": "",
                "overlijden_j": "",
                "overlijden_m": "",
                "overlijden_d": "",
                "opmerkingen": "",
            }
        ]
    )
    parsed = parse_persoon_import_file("clear.csv", payload)
    assert parsed["changes"][0]["value"] == ""


def test_xlsx_roundtrip():
    wb = Workbook()
    ws = wb.active
    ws.title = WORKSHEET_NAME
    for idx, key in enumerate(IMPORT_COLUMN_KEYS, start=1):
        ws.cell(row=1, column=idx, value=key)
    ws.cell(row=2, column=1, value=99)
    ws.cell(row=2, column=5, value=1720)
    buf = io.BytesIO()
    wb.save(buf)
    parsed = parse_persoon_import_file("upload.xlsx", buf.getvalue())
    assert parsed["person_count"] == 1
    assert any(c["field"] == "geboortejaar" and c["value"] == "1720" for c in parsed["changes"])
