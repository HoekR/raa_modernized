"""Excel/CSV import for editorial persoon werklijst (fixed schema)."""

from __future__ import annotations

import csv
import io
from dataclasses import dataclass
from typing import Any, BinaryIO

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy.orm import Session

from raa_api.editorial_batch import apply_batch_changes, fetch_batch_rows

CLEAR_MARKER = "-"

WORKSHEET_NAME = "werklijst"
INSTRUCTIONS_SHEET = "uitleg"
TEMPLATE_FILENAME = "raa_persoon_werklijst.xlsx"
MAX_IMPORT_ROWS = 500


@dataclass(frozen=True)
class ImportColumn:
    key: str
    field: str | None  # None for persoon_id column
    hint: str


PERSOON_IMPORT_COLUMNS: tuple[ImportColumn, ...] = (
    ImportColumn("persoon_id", None, "Verplicht. Bestaand RAA persoon-id."),
    ImportColumn("geslachtsnaam", "geslachtsnaam", "Achternaam."),
    ImportColumn("voornaam", "voornaam", "Voornaam."),
    ImportColumn("tussenvoegsel", "tussenvoegsel", "Tussenvoegsel."),
    ImportColumn("geboorte_j", "geboortejaar", "Geboortejaar (1400–1920)."),
    ImportColumn("geboorte_m", "geboortemaand", "Geboortemaand 1–12 (optioneel)."),
    ImportColumn("geboorte_d", "geboortedag", "Geboortedag 1–31 (optioneel, vereist m)."),
    ImportColumn("overlijden_j", "overlijdensjaar", "Overlijdensjaar (1400–1920)."),
    ImportColumn("overlijden_m", "overlijdensmaand", "Overlijdensmaand 1–12 (optioneel)."),
    ImportColumn("overlijden_d", "overlijdensdag", "Overlijdensdag 1–31 (optioneel, vereist m)."),
    ImportColumn("opmerkingen", "opmerkingen", "Vrije tekst."),
)

IMPORT_COLUMN_KEYS: tuple[str, ...] = tuple(c.key for c in PERSOON_IMPORT_COLUMNS)


def _instruction_lines() -> list[str]:
    lines = [
        "RAA redactie — persoon werklijst",
        "",
        "1. Bewerk alleen kolommen die u wilt wijzigen; lege cellen worden overgeslagen.",
        f"2. Gebruik '{CLEAR_MARKER}' om een veld expliciet leeg te maken.",
        "3. Datums: exact j / m / d — maand en dag zijn optioneel.",
        "4. Sla op als .xlsx (dit bestand) of UTF-8 CSV met dezelfde kolomkoppen.",
        "5. Max 500 datarijen per import.",
        "",
        "Kolommen:",
    ]
    lines.extend(f"  • {col.key}: {col.hint}" for col in PERSOON_IMPORT_COLUMNS)
    return lines


def _style_header(ws: Worksheet) -> None:
    header_font = Font(bold=True)
    fill = PatternFill("solid", fgColor="E8F0F4")
    for idx, col in enumerate(PERSOON_IMPORT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=idx, value=col.key)
        cell.font = header_font
        cell.fill = fill
        width = 14 if col.key == "persoon_id" else 16
        if col.key == "opmerkingen":
            width = 28
        ws.column_dimensions[cell.column_letter].width = width
    ws.freeze_panes = "A2"


def _prefill_rows(db: Session, entity_ids: list[int]) -> list[dict[str, str]]:
    if not entity_ids:
        return []
    batch = fetch_batch_rows(db, "persoon", entity_ids)
    rows: list[dict[str, str]] = []
    key_for_field = {c.field: c.key for c in PERSOON_IMPORT_COLUMNS if c.field}
    for row in batch["rows"]:
        out: dict[str, str] = {"persoon_id": str(row["id"])}
        for field, col_key in key_for_field.items():
            out[col_key] = row["fields"].get(field, {}).get("effective", "")
        rows.append(out)
    return rows


def build_persoon_template_xlsx(
    db: Session | None = None,
    *,
    entity_ids: list[int] | None = None,
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = WORKSHEET_NAME
    _style_header(ws)

    prefill = _prefill_rows(db, entity_ids or []) if db is not None else []
    for row_idx, data in enumerate(prefill, start=2):
        for col_idx, col in enumerate(PERSOON_IMPORT_COLUMNS, start=1):
            ws.cell(row=row_idx, column=col_idx, value=data.get(col.key, ""))

    info = wb.create_sheet(INSTRUCTIONS_SHEET)
    for i, line in enumerate(_instruction_lines(), start=1):
        info.cell(row=i, column=1, value=line)
    info.column_dimensions["A"].width = 90

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _normalize_cell(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.lower() in {"none", "nan"}:
        return ""
    return text


def _parse_row_dict(
    raw: dict[str, Any], row_number: int
) -> tuple[int | None, list[dict[str, Any]], list[dict[str, Any]]]:
    changes: list[dict[str, Any]] = []
    errors: list[dict[str, Any]] = []

    try:
        pid_raw = _normalize_cell(raw.get("persoon_id"))
        if not pid_raw:
            errors.append({"row": row_number, "column": "persoon_id", "error": "persoon_id ontbreekt"})
            return None, changes, errors
        persoon_id = int(float(pid_raw.replace(".0", "")))
        if persoon_id <= 0:
            raise ValueError("ongeldig id")
    except ValueError:
        errors.append(
            {
                "row": row_number,
                "column": "persoon_id",
                "error": f"ongeldig persoon_id: {raw.get('persoon_id')!r}",
            }
        )
        return None, changes, errors

    for col in PERSOON_IMPORT_COLUMNS:
        if col.field is None:
            continue
        cell = _normalize_cell(raw.get(col.key))
        if not cell:
            continue
        value = "" if cell == CLEAR_MARKER else cell
        changes.append(
            {
                "entity_type": "persoon",
                "entity_id": persoon_id,
                "field": col.field,
                "value": value,
            }
        )

    return persoon_id, changes, errors


def _validate_headers(headers: list[str]) -> list[dict[str, Any]]:
    normalized = [_normalize_cell(h) for h in headers]
    if normalized != list(IMPORT_COLUMN_KEYS):
        return [
            {
                "row": 1,
                "column": "headers",
                "error": (
                    "Kolomkoppen komen niet overeen met het sjabloon. "
                    f"Verwacht: {', '.join(IMPORT_COLUMN_KEYS)}"
                ),
            }
        ]
    return []


def _rows_from_csv(data: bytes) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    text = data.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        return [], [{"row": 1, "column": "headers", "error": "Geen kolomkoppen gevonden"}]
    errors = _validate_headers(list(reader.fieldnames))
    if errors:
        return [], errors
    return list(reader), []


def _rows_from_xlsx(data: bytes) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb[WORKSHEET_NAME] if WORKSHEET_NAME in wb.sheetnames else wb.active
    iterator = ws.iter_rows(values_only=True)
    try:
        header_row = next(iterator)
    except StopIteration:
        return [], [{"row": 1, "column": "headers", "error": "Leeg werkblad"}]
    headers = [_normalize_cell(h) for h in header_row]
    errors = _validate_headers(headers)
    if errors:
        return [], errors

    rows: list[dict[str, Any]] = []
    for values in iterator:
        if not values or all(_normalize_cell(v) == "" for v in values):
            continue
        row_dict = {
            headers[i]: values[i] if i < len(values) else None for i in range(len(headers))
        }
        rows.append(row_dict)
    return rows, []


def parse_persoon_import_file(filename: str, data: bytes) -> dict[str, Any]:
    lower = filename.lower()
    if lower.endswith(".csv"):
        raw_rows, header_errors = _rows_from_csv(data)
    elif lower.endswith((".xlsx", ".xlsm")):
        raw_rows, header_errors = _rows_from_xlsx(data)
    else:
        return {
            "parse_errors": [{"row": 0, "column": "file", "error": "Alleen .xlsx of .csv toegestaan"}],
            "changes": [],
            "rows_parsed": 0,
        }

    if header_errors:
        return {"parse_errors": header_errors, "changes": [], "rows_parsed": 0}

    if len(raw_rows) > MAX_IMPORT_ROWS:
        return {
            "parse_errors": [
                {"row": 0, "column": "file", "error": f"Max {MAX_IMPORT_ROWS} datarijen per import"}
            ],
            "changes": [],
            "rows_parsed": 0,
        }

    parse_errors: list[dict[str, Any]] = []
    changes: list[dict[str, Any]] = []
    seen_ids: set[int] = set()

    for idx, raw in enumerate(raw_rows, start=2):
        persoon_id, row_changes, row_errors = _parse_row_dict(raw, idx)
        parse_errors.extend(row_errors)
        if persoon_id is not None:
            if persoon_id in seen_ids:
                parse_errors.append(
                    {"row": idx, "column": "persoon_id", "error": f"dubbel persoon_id {persoon_id}"}
                )
            else:
                seen_ids.add(persoon_id)
        changes.extend(row_changes)

    return {
        "parse_errors": parse_errors,
        "changes": changes,
        "rows_parsed": len(raw_rows),
        "person_count": len(seen_ids),
        "change_count": len(changes),
    }


def import_persoon_file(
    db: Session,
    *,
    filename: str,
    data: bytes,
    editor_id: str,
    note: str | None = None,
    dry_run: bool = False,
) -> dict[str, Any]:
    parsed = parse_persoon_import_file(filename, data)
    if parsed["parse_errors"]:
        return {**parsed, "result": None, "dry_run": dry_run}

    if not parsed["changes"]:
        return {
            **parsed,
            "result": None,
            "dry_run": dry_run,
            "parse_errors": parsed["parse_errors"]
            + [{"row": 0, "column": "file", "error": "Geen wijzigingen gevonden (alle cellen leeg)"}],
        }

    result = apply_batch_changes(
        db,
        changes=parsed["changes"],
        editor_id=editor_id,
        note=note,
        dry_run=dry_run,
    )
    return {**parsed, "result": result, "dry_run": dry_run}


def read_upload_stream(stream: BinaryIO) -> bytes:
    return stream.read()
